#!/usr/bin/env python
# coding: utf-8

# # WEEK 4
# #### the following are a series of various was to open and deal with different file types.
# 

# In[1]:


#4-1
# A simple example of reading data from a .csv file with Python
# using the "csv" library.


# In[2]:


# importing the library to work with csv
import csv

# opening the document with the open function to "read"
source_file = open("202009CitibikeTripdataExample.csv","r")

#passing the source_file through the csv reader, DictReader, and storing it in a variable called "citibike_reader"
citibike_reader = csv.DictReader(source_file)

#printing the headings of each column in the dataset using the .fieldnames function.
print(citibike_reader.fieldnames)

# let's just print out the first 5 rows
for i in range(0,5):
    print (next(citibike_reader))


# #### in the above cell, 4-1, I used the csv function to open a csv (comma separated document) and print first the headings of each row and then the attributes of the first few rows.

# In[3]:


#4-3: 
# A simple example of reading data from a .tsv file with Python, using
# the `csv` library. The source data was downloaded as a .tsv file


# In[4]:


import csv

txt_source_file = open("ShugermanProsecutorPoliticians-SupremeCourtJustices.txt","r")

# pass our tsv_source_file as an ingredient to the the csv library's DictReader
# "recipe" and store the result in a variable called `politicians_reader`
# add the "delimiter" parameter and specify the tab character, "\t"
politicians_reader = csv.DictReader(txt_source_file, delimiter='\t')
# the DictReader function has added useful information to our data,
# like a label that shows us all the values in the first or "header" row
print(politicians_reader.fieldnames)
# we'll use the `next()` function to print just the first row of data
print (next(politicians_reader))


# #### 4-3 we used the csv method again but this time the file was separated by tabs instead of commas. we are doing the same as 4-1 but we change the delimiter to a tab. we print the heading along with the first row.

# In[5]:


#4-4
# An example of reading data from an .xlsx file with Python, using the "openpyxl"


# In[6]:


from openpyxl import load_workbook

# import the `csv` library, to create our output file
import csv

source_workbook = load_workbook(filename = 'fredgraph.xlsx')

# an .xlsx workbook can have multiple sheets
# print their names here for reference
print(source_workbook.sheetnames)
# loop through the worksheets in `source_workbook`
for sheet_num, sheet_name in enumerate(source_workbook.sheetnames):

    # create a variable that points to the current worksheet by
    # passing the current value of `sheet_name` to `source_workbook`
    current_sheet = source_workbook[sheet_name]

    # print `sheet_name`, just to see what it is
    print(sheet_name)

    # create an output file called "xlsx_"+sheet_name
    output_file = open("xlsx_"+sheet_name+".csv","w")

    # use this csv library's "writer" recipe to easily write rows of data
    # to `output_file`, instead of reading data *from* it
    output_writer = csv.writer(output_file)

    # loop through every row in our sheet
    for row in current_sheet.iter_rows():

        # we'll create an empty list where we'll put the actual
        # values of the cells in each row
        row_cells = []

        # for every cell (or column) in each row....
        for cell in row:

            # let's print what's in here, just to see how the code sees it
            print(cell, cell.value)

            # add the values to the end of our list with the `append()` method
            row_cells.append(cell.value)

        # write our newly (re)constructed data row to the output file
        output_writer.writerow(row_cells)
# officially close the `.csv` file we just wrote all that data to
output_file.close()


# #### 4-4 is code that takes an excel file, often with many sheets and converts it into single CSV files. the output was very long, so I wanted to remove the print function, but then I decided that it feels good to see the output so I put it back. 

# In[7]:


#4-6
# A simple example of reading data from a .xls file with Python
# using the "xrld" library. First, pip install the xlrd library:


# In[8]:


get_ipython().system('pip install xlrd')
import xlrd

# import the `csv` library, to create our output file
import csv

source_workbook = xlrd.open_workbook("fredgraph.xls")
# an `.xls` workbook can have multiple sheets
for sheet_name in source_workbook.sheet_names():

    # create a variable that points to the current worksheet by
    # passing the current value of `sheet_name` to the `sheet_by_name` recipe
    current_sheet = source_workbook.sheet_by_name(sheet_name)

    # print `sheet_name`, just to see what it is
    print(sheet_name)

    # create "xls_"+sheet_name+".csv" as an output file for the current sheet
    output_file = open("xls_"+sheet_name+".csv","w")

    # use the `csv` library's "writer" recipe to easily write rows of data
    # to `output_file`, instead of reading data *from* it
    output_writer = csv.writer(output_file)

    # now, we need to loop through every row in our sheet
    for row_num, row in enumerate(current_sheet.get_rows()):

        # each row is already a list, but we need to use the `row_value()`
        # method to access them
        # then we can use the `writerow` recipe to write them
        # directly to our output file
        output_writer.writerow(current_sheet.row_values(row_num))

    # officially close the `.csv` file we just wrote all that data to
    output_file.close()
# files.download("xls_"+sheet_name+".csv")


# #### this code reads files from .xls format with python using the xrld library. The first time I ran it I got an error and then I realized that I forgot to install the library.

# In[9]:


#4-7 fixed width parsing
# An example of reading data from a fixed-width file with Python.


# In[10]:


import csv

# imported_file = drive.CreateFile({'id': file_id}) # creating an accessible copy of the shared data file
# print(imported_file['title'])  # it should print the title of desired file
# imported_file.GetContentFile(imported_file['title']) # refer to it in this notebook by the same name as it has in Drive
filename = "ghcnd-stations"

# reading from a basic text file doesn't require any special libraries
# so we'll just open the file in read format ("r") as usual
source_file = open(filename+".txt", "r")

# the built-in "readlines()" method does just what you'd think:
# it reads in a text file and converts it to a list of lines
stations_list = source_file.readlines()

# as usual, we'll create an output file to write to
output_file = open(filename+".csv","w")

# and we'll use the `csv` library to create a "writer" that gives us handy
# "recipe" functions for creating our new file in csv format
output_writer = csv.writer(output_file)
# create the header list
headers = ["ID","LATITUDE","LONGITUDE","ELEVATION","STATE","NAME","GSN_FLAG",
 "HCNCRN_FLAG","WMO_ID"]

# write our headers to the output file
output_writer.writerow(headers)
# loop through each line of our file (multiple "sheets" are not possible)
for line in stations_list:
    # create an empty list, to which we'll append each set of characters that
    # makes up a given "column" of data
    new_row = []
    # ID: positions 1-11
    new_row.append(line[0:11])
    # LATITUDE: positions 13-20
    new_row.append(line[12:20])
    # LONGITUDE: positions 22-30
    new_row.append(line[21:30])
    # ELEVATION: positions 32-37
    new_row.append(line[31:37])
    # STATE: positions 39-40
    new_row.append(line[38:40])
    # NAME: positions 42-71
    new_row.append(line[41:71])
    # GSN_FLAG: positions 73-75
    new_row.append(line[72:75])
    # HCNCRN_FLAG: positions 77-79
    new_row.append(line[76:79])
    # WMO_ID: positions 81-85
    new_row.append(line[80:85])

    # now all that's left is to use the
    # `writerow` function to write new_row to our output file
    output_writer.writerow(new_row)
# officially close the `.csv` file we just wrote all that data to
output_file.close()
print(f"Successfully converted {filename}.txt to {filename}.csv")


# #### A fixed width file is file that each column has a fixed width, instead of having a delimiter. The way to separate each attribute, is by indexing the amount of figures in each column. The main chunk of this code is writing out the index of every column. The new file is rewritten as a CSV. There was no output, and I was able to see that it went through because I have a new file now that is a CSV. I also added a line to print when the conversion is successful. 

# In[11]:


#4-15
# A simple example of reading data from a .json file with Python,
# using the built-in "json" library.


# In[12]:


import json

# import the `csv` library, to create our output file
import csv

# imported_file = drive.CreateFile({'id': file_id}) # creating an accessible copy of the shared data file
# print(imported_file['title'])  # it should print the title of desired file
# imported_file.GetContentFile(imported_file['title']) # refer to it in this notebook by the same name as it has in Drive
# choose a filename
filename = "U6_FRED_data"
# open the file in read format ("r") as usual
json_source_file = open(filename+".json","r")

# pass the `json_source_file` as an ingredient to the json library's `load()`
# method and store the result in a variable called `json_data`
json_data = json.load(json_source_file)
# create our output file, naming it "json_"+filename
output_file = open("json_"+filename+".csv","w")

# use the `csv` library's "writer" recipe to easily write rows of data
# to `output_file`, instead of reading data *from* it
output_writer = csv.writer(output_file)
# grab the first element (at position "0"), and use its keys as the column headers
output_writer.writerow(list(json_data["observations"][0].keys()))
for obj in json_data["observations"]:

    # we'll create an empty list where we'll put the actual values of each object
    obj_values = []

    # for every `key` (which will become a column), in each object
    for key, value in obj.items():

        # let's print what's in here, just to see how the code sees it
        print(key,value)

        # add the values to our list
        obj_values.append(value)

    # now we've got the whole row, write the data to our output file
    output_writer.writerow(obj_values)
# officially close the `.csv` file we just wrote all that data to
output_file.close()


# #### 4-15 reads a .json file in python with the json library

# In[13]:


#4-16
# A basic example of reading data from a .pdf file with Python,
# using `pdf2image` to convert it to images, and then using the
# openCV and `tesseract` libraries to extract the text


# In[14]:


get_ipython().system('pip install pdf2image')
get_ipython().system('pip install pytesseract')
import os

# we'll import the `convert_from_path` "chapter" of the `pdf2image` library
from pdf2image import convert_from_path

# the built-in `glob`library offers a handy way to loop through all the files
# in a folder that have a certain file extension, for example
import glob

# `cv2` is the actual library name for `openCV`
import cv2

# and of course, we need our Python library for interfacing
# with the tesseract OCR process
import pytesseract
# we'll use the pdf name to name both our generated images and text files
pdf_name = "SafetyNet"

# our source pdf is in the same folder as our Python script
pdf_source_file = pdf_name+".pdf"

# as long as a folder with the same name as the pdf does not already exist
if os.path.isdir(pdf_name) == False:
    # create a new folder with that name
    target_folder = os.mkdir(pdf_name)

pages = convert_from_path(pdf_source_file, 300)
# loop through all the converted pages, enumerating them so that the page
# number can be used to label the resulting images
for page_num, page in enumerate(pages):
    # create unique filenames for each page image, combining the
    # folder name and the page number
    filename = os.path.join(pdf_name,"p"+str(page_num)+".png")
    # save the image of the page in system
    page.save(filename, 'PNG')
# next, go through all the files in the folder that end in `.png`
for img_file in glob.glob(os.path.join(pdf_name, '*.png')):
    # replace the slash in the image's filename with a dot
    temp_name = img_file.replace("/",".")
    # pull the unique page name (e.g. `p2`) from the `temp_name`
    text_filename = temp_name.split(".")[1]
    # now! create a new, writable file, also in our target folder, that
    # has the same name as the image, but is a `.txt` file
    output_file = open(os.path.join(pdf_name,text_filename+".txt"), "w")
    # use the `cv2` library to interpret our image
    img = cv2.imread(img_file)
    # create a new variable to hold the results of using pytesseract's
    # `image_to_string()` function, which will do just that
    converted_text = pytesseract.image_to_string(img)
    # write our extracted text to our output file
    output_file.write(converted_text)
    # close the output file
    output_file.close()


# #### This example was my favorite, we took a pdf file converted it to images and extracted the text form the images. I had a few hiccups installing the proper libraries in the correct folders... I was so intrigued so I opened the new files a bunch are images containing a lot of text and the other are files of the text that was extracted from it.

# #### I used all the code straight from the book's Github. I modified a drop and mentioned it whenever I did.

# In[ ]:




